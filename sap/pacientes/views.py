from django.forms import modelform_factory
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.template import loader
from openpyxl.workbook import Workbook

from .forms import PacienteFormulario
from .models import Paciente

def agregar_paciente(request):
    if request.method == 'GET':
        formulario = PacienteFormulario()
    elif request.method == 'POST':
        formulario = PacienteFormulario(request.POST)
        if formulario.is_valid():
            formulario.save()
            return redirect('inicio')  # Ajusta esta redirección según tu nombre de ruta

    contexto = {'formulario': formulario}
    return render(request, 'pacientes/agregar.html', contexto)

def modificar_paciente(request, id):
    pagina = loader.get_template('pacientes/modificar.html')
    paciente = get_object_or_404(Paciente, pk=id)
    if request.method == 'GET':
        formulario = PacienteFormulario(instance=paciente)
    elif request.method == 'POST':
        formulario = PacienteFormulario(request.POST, instance=paciente)
        if formulario.is_valid():
            formulario.save()
            return redirect('inicio')
    datos = {'formulario': formulario}
    return HttpResponse(pagina.render(datos, request))

def ver_paciente(request, id):
    paciente = get_object_or_404(Paciente, pk=id)
    datos = {'paciente': paciente}
    pagina = loader.get_template('pacientes/ver.html')
    return HttpResponse(pagina.render(datos, request))

def eliminar_paciente(request, id):
    paciente = get_object_or_404(Paciente, pk=id)
    if paciente:
        paciente.delete()
        return redirect('inicio')


def generar_reporte(request, *args, **kwargs):
    # Obtener todos los pacientes de la base de datos ordenados por apellido y nombre
    pacientes = Paciente.objects.order_by('apellido', 'nombre')

    # Crear el libro de trabajo
    wb = Workbook()
    ws = wb.active

    # Encabezados del reporte
    ws['B1'] = 'REPORTE DE PACIENTES'
    ws.merge_cells('B1:F1')
    ws['B3'] = 'ID'
    ws['C3'] = 'Nombre'
    ws['D3'] = 'Apellido'
    ws['E3'] = 'Email'


    cont = 4
    # Recorrer los pacientes y escribir cada uno de los datos en las celdas
    for paciente in pacientes:
        ws.cell(row=cont, column=2).value = paciente.id
        ws.cell(row=cont, column=3).value = paciente.nombre
        ws.cell(row=cont, column=4).value = paciente.apellido
        ws.cell(row=cont, column=5).value = paciente.email

        cont += 1

    # Establecer el nombre del archivo
    nombre_archivo = "ReportePacientesExcel.xlsx"
    # Definir el tipo de respuesta como un archivo de Microsoft Excel
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response